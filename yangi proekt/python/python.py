import streamlit as st
import pandas as pd
from google import genai
from google.genai import types
from PIL import Image
import io

# 1. Sahifa dizayni va sarlavhasi
st.set_page_config(page_title="Pro Skaner v2", page_icon="📊", layout="wide")

st.markdown("""
    <style>
    .main-title { font-size:32px; font-weight:bold; color:#1E3A8A; text-align:center; margin-bottom:20px; }
    .stButton>button { width: 100%; background-color: #10B981; color: white; font-weight: bold; }
    </style>
    """, unsafe_allow_html=True)

st.markdown("<div class='main-title'>📊 Professional Rasm -> Excel Skaner Tizimi</div>", unsafe_allow_html=True)
st.write("Bir yoki bir nechta daftar rasmlarini yuklang, tizim ularni bitta umumiy jadvalga yig'adi.")

# --- API Kalit ---
GEMINI_API_KEY = "AIzaSyBHQC06bNnoaWRx88ZYzdAnY9eQyLYuZ2w"

# --- 1. Ko'p sonli rasmlarni bir vaqtda yuklash (Batch Processing) ---
uploaded_files = st.file_uploader("Rasmlarni tanlang (Bir nechta yuklash mumkin)...", type=["jpg", "jpeg", "png"], accept_multiple_files=True)

if uploaded_files:
    st.info(f"📁 Jami {len(uploaded_files)} ta rasm yuklandi.")
    
    # Rasmlarni yonma-yon kichik ko'rinishda chiqarish
    cols = st.columns(min(len(uploaded_files), 5))
    for idx, file in enumerate(uploaded_files):
        with cols[idx % 5]:
            img = Image.open(file)
            st.image(img, caption=file.name, use_container_width=True)

    # Skanerlash tugmasi
    if st.button("🚀 Barcha rasmlarni skanerlash va birlashtirish"):
        if GEMINI_API_KEY == "SIZNING_GEMINI_API_KALITINGIZ" or not GEMINI_API_KEY:
            st.error("🚨 Iltimos, kod ichiga o'zingizning Google Gemini API kalitingizni kiriting!")
        else:
            all_records = []
            progress_bar = st.progress(0)
            
            for index, file in enumerate(uploaded_files):
                st.write(f"🔍 {file.name} tahlil qilinmoqda...")
                try:
                    image = Image.open(file)
                    img_byte_arr = io.BytesIO()
                    image.save(img_byte_arr, format=image.format if image.format else 'JPEG')
                    img_bytes = img_byte_arr.getvalue()
                    
                    client = genai.Client(api_key=GEMINI_API_KEY)
                    image_part = types.Part.from_bytes(
                        data=img_bytes,
                        mime_type=f"image/{image.format.lower() if image.format else 'jpeg'}",
                    )
                    
                    prompt = """Rasmda qo'lda yozilgan kodlar ro'yxati bor. Har bir kodni ikki qismga ajrat.
                    Masalan: '06-60-01-0613-20 W' kodi uchun: Artikul = '06-60-01-0613', Son = '20'.
                    Oxiridagi 'W' harfini mutlaqo tashlab yubor. Natijani faqat CSV formatda qaytar,
                    ustunlar nomi 'Artikul,Son' bo'lsin. Hech qanday ortiqcha izoh yoki matn yozma."""
                    
                    response = client.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=[image_part, prompt]
                    )
                    
                    csv_data = response.text.replace("```csv", "").replace("```", "").strip()
                    lines = [line.split(',') for line in csv_data.split('\n') if ',' in line]
                    
                    if len(lines) > 1:
                        for row in lines[1:]:
                            if len(row) == 2:
                                all_records.append({"Artikul": row[0].strip(), "Son": row[1].strip()})
                                
                except Exception as e:
                    st.error(f"❌ {file.name} faylida xatolik: {str(e)}")
                
                # Progress panelini yangilash
                progress_bar.progress((index + 1) / len(uploaded_files))
            
            if all_records:
                st.session_state['scanned_df'] = pd.DataFrame(all_records)
                st.success("🎉 Barcha rasmlar muvaffaqiyatli skanerlandi!")

# --- 2. Skanerlash natijasi va Tahrirlash oyna qismi ---
if 'scanned_df' in st.session_state:
    st.subheader("📝 2. Ma'lumotlarni tekshirish va tahrirlash")
    st.write("Jadvaldagi istalgan katakchani ikki marta bosib, xatolarni qo'lda to'g'rilashingiz mumkin.")
    
    # Brauzerda tahrirlanuvchi jadval (Editable Table)
    edited_df = st.data_editor(st.session_state['scanned_df'], num_rows="dynamic", height=400, use_container_width=True)
    
    # --- 3. Dublikatlarni tekshirish (Lekin sonlarni qo'shmaslik) ---
    # Takrorlangan artikullarni topamiz (lekin o'chirib tashlamaymiz yoki qo'shmaymiz)
    duplicate_artikuls = edited_df[edited_df.duplicated(subset=['Artikul'], keep=False)]['Artikul'].unique()
    
    if len(duplicate_artikuls) > 0:
        st.warning(f"⚠️ Diqqat! Tizimda quyidagi artikullar bir necha marta takrorlangan (dublikat): {', '.join(duplicate_artikuls)}. Ular Excelga qo'shilmasdan, alohida qator bo'lib tushadi.")
    
    # --- 4. Excel fayl dizaynini chiroyli qilish (Styling) ---
    st.subheader("📥 3. Tayyor faylni yuklab olish")
    
    excel_buffer = io.BytesIO()
    # Excel faylga chiroyli dizayn berish (openpyxl orqali)
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        edited_df.to_excel(writer, index=False, sheet_name="Artikullar")
        
        # Dizayn elementlarini qo'shish
        workbook = writer.book
        worksheet = writer.sheets["Artikullar"]
        
        # Sarlavha dizayni (To'q ko'k va oq yozuv)
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Chegaralar (Borders)
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        # Sarlavhani bezash
        for col_num in range(1, len(edited_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
        # Ustun kengliklarini raqamlarga moslab avtomatik kengaytirish
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
            # Ma'lumot qatorlariga chegara qo'shish va markazlashtirish
            for cell in col[1:]:
                cell.alignment = center_alignment
                cell.border = thin_border

    excel_data = excel_buffer.getvalue()
    
    st.download_button(
        label="📥 Dizaynli Excel faylni yuklab olish",
        data=excel_data,
        file_name="skanerlangan_artikullar_pro.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
